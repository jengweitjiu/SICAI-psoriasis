#!/usr/bin/env python3
"""
Part_J_Populate_Supp_Tables.py
────────────────────────────────
Reads actual CSV results from SICAI pipeline (Parts D–H) and populates
the Supplementary Tables xlsx template with real data.

Run on Colab after uploading SICAI_Supplementary_Tables.xlsx to Drive.
"""

import os, glob, sys
import pandas as pd
import numpy as np

# ── pip install if needed ──
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    os.system("pip install openpyxl -q")
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ═══════════════════════════════════════════════════════════════
# CONFIG — adjust paths to match your Drive layout
# ═══════════════════════════════════════════════════════════════
BASE = "/content/drive/MyDrive/Fibroblast_Atlas"
RESULTS_DIRS = [
    f"{BASE}/results_colab",
    f"{BASE}/results_colab/manuscript_figures",
    f"{BASE}/results",
    f"{BASE}/SICAI_results",
    BASE,
]
TEMPLATE = f"{BASE}/submission/SICAI_Supplementary_Tables.xlsx"
OUTPUT   = f"{BASE}/submission/SICAI_Supplementary_Tables_FINAL.xlsx"

# ═══════════════════════════════════════════════════════════════
# HELPER: find CSV by pattern
# ═══════════════════════════════════════════════════════════════
def find_csv(patterns, label=""):
    """Search RESULTS_DIRS for CSV matching any pattern."""
    for d in RESULTS_DIRS:
        if not os.path.isdir(d):
            continue
        for f in os.listdir(d):
            fl = f.lower()
            if not fl.endswith('.csv'):
                continue
            for pat in patterns:
                if pat.lower() in fl:
                    path = os.path.join(d, f)
                    print(f"  ✅ {label}: {path}")
                    return path
    # Deep search
    for d in RESULTS_DIRS[:2]:
        for root, dirs, files in os.walk(d):
            for f in files:
                fl = f.lower()
                if not fl.endswith('.csv'):
                    continue
                for pat in patterns:
                    if pat.lower() in fl:
                        path = os.path.join(root, f)
                        print(f"  ✅ {label} (deep): {path}")
                        return path
    print(f"  ❌ {label}: NOT FOUND (patterns: {patterns})")
    return None

# ═══════════════════════════════════════════════════════════════
# STYLING HELPERS
# ═══════════════════════════════════════════════════════════════
data_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

def style_cell(ws, row, col, value, num_fmt=None, align='center'):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = data_font
    cell.border = thin_border
    cell.alignment = Alignment(horizontal=align, vertical='center')
    if num_fmt:
        cell.number_format = num_fmt
    return cell

# ═══════════════════════════════════════════════════════════════
# LOCATE ALL SOURCE CSVs
# ═══════════════════════════════════════════════════════════════
print("=" * 60)
print("LOCATING SOURCE CSVs")
print("=" * 60)

# S1 sources
sicai_csv    = find_csv(['sicai_result', 'sicai_104', 'sicai_composite'], "S1: SICAI results")
cs_csv       = find_csv(['cs_matrix', 'coupling_strength'], "S1: CS matrix")
csp_csv      = find_csv(['csp_matrix', 'coupling_specific'], "S1: CSp matrix")
pasi_corr    = find_csv(['pasi_corr', 'severity_corr', 'cs_pasi'], "S1: PASI correlations")

# S2 sources
lr_csv       = find_csv(['table_s2', 'lr_spatial', 'lr_coexp', 'ligand_receptor'], "S2: L-R results")

# S3 sources
sample_csv   = find_csv(['sample_metric', 'per_sample', 'gsd_metric', 'basin_metric'], "S3: Sample metrics")
gsd_csv      = find_csv(['gsd_result', 'gsd_score', 'attractor', 'pca_score'], "S3: GSD scores")

# S4 — keep template (curated gene lists don't come from pipeline)
print(f"  ℹ️  S4: Immune signatures — keeping curated template")

# ═══════════════════════════════════════════════════════════════
# LOAD TEMPLATE
# ═══════════════════════════════════════════════════════════════
print(f"\n{'='*60}")
print(f"LOADING TEMPLATE: {TEMPLATE}")

if not os.path.exists(TEMPLATE):
    # Try alternative locations
    alt = glob.glob(f"{BASE}/**/SICAI_Supplementary*.xlsx", recursive=True)
    if alt:
        TEMPLATE = alt[0]
        print(f"  → Found at: {TEMPLATE}")
    else:
        print(f"  ❌ Template not found! Upload SICAI_Supplementary_Tables.xlsx to {BASE}/submission/")
        sys.exit(1)

wb = load_workbook(TEMPLATE)
replaced = {}

# ═══════════════════════════════════════════════════════════════
# TABLE S1: SICAI RESULTS
# ═══════════════════════════════════════════════════════════════
ws1 = wb["Table S1 - SICAI Results"]

def populate_s1():
    """Replace S1 with actual SICAI pipeline data."""
    
    # Strategy A: Single comprehensive SICAI CSV
    if sicai_csv:
        df = pd.read_csv(sicai_csv)
        print(f"\n  S1 source: {sicai_csv}")
        print(f"  Columns: {list(df.columns)}")
        print(f"  Shape: {df.shape}")
        
        # Map columns flexibly
        col_map = {}
        for col in df.columns:
            cl = col.lower().strip()
            if cl in ['fibroblast', 'fibroblast_subtype', 'fib_subtype', 'stromal']:
                col_map['fibroblast'] = col
            elif cl in ['immune', 'immune_population', 'immune_type', 'immune_cell']:
                col_map['immune'] = col
            elif cl in ['cs', 'coupling_strength', 'cs_rho', 'spearman_rho']:
                col_map['cs'] = col
            elif cl in ['csp', 'coupling_specificity', 'csp_sigma']:
                col_map['csp'] = col
            elif cl in ['as', 'attractor_stability', 'as_score', 'stability']:
                col_map['as'] = col
            elif cl in ['sicai', 'sicai_score', 'composite']:
                col_map['sicai'] = col
            elif 'pasi' in cl and ('corr' in cl or 'rho' in cl or 'r' == cl[-1]):
                col_map['pasi_r'] = col
            elif 'pasi' in cl and 'p' in cl:
                col_map['pasi_p'] = col
        
        print(f"  Mapped columns: {col_map}")
        
        if 'fibroblast' in col_map and 'immune' in col_map:
            # Clear old data rows
            for r in range(5, ws1.max_row + 1):
                for c in range(1, 19):
                    ws1.cell(row=r, column=c).value = None
            
            row = 5
            for idx, r_data in df.iterrows():
                fb = r_data[col_map['fibroblast']]
                im = r_data[col_map['immune']]
                cs_val = r_data.get(col_map.get('cs', ''), None)
                csp_val = r_data.get(col_map.get('csp', ''), None)
                as_val = r_data.get(col_map.get('as', ''), None)
                sicai_val = r_data.get(col_map.get('sicai', ''), None)
                
                style_cell(ws1, row, 1, idx + 1)
                style_cell(ws1, row, 2, str(fb), align='left')
                style_cell(ws1, row, 3, str(im))
                
                if cs_val is not None and pd.notna(cs_val):
                    style_cell(ws1, row, 4, float(cs_val), '0.000')
                if csp_val is not None and pd.notna(csp_val):
                    style_cell(ws1, row, 5, float(csp_val), '0.000')
                if as_val is not None and pd.notna(as_val):
                    style_cell(ws1, row, 6, float(as_val), '0.000')
                if sicai_val is not None and pd.notna(sicai_val):
                    style_cell(ws1, row, 7, float(sicai_val), '0.0000')
                
                # PASI correlation if available
                pasi_r_val = r_data.get(col_map.get('pasi_r', ''), None)
                if pasi_r_val is not None and pd.notna(pasi_r_val):
                    style_cell(ws1, row, 11, float(pasi_r_val), '0.000')
                pasi_p_val = r_data.get(col_map.get('pasi_p', ''), None)
                if pasi_p_val is not None and pd.notna(pasi_p_val):
                    style_cell(ws1, row, 12, float(pasi_p_val), '0.00E+00')
                
                # Copy any additional columns present
                for extra_col in df.columns:
                    ecl = extra_col.lower()
                    if 'healthy' in ecl or 'cs_healthy' in ecl:
                        v = r_data[extra_col]
                        if pd.notna(v): style_cell(ws1, row, 8, float(v), '0.000')
                    elif 'non_lesional' in ecl or 'nonlesional' in ecl or 'cs_nl' in ecl:
                        v = r_data[extra_col]
                        if pd.notna(v): style_cell(ws1, row, 9, float(v), '0.000')
                    elif 'lesional' in ecl and 'non' not in ecl or 'cs_les' in ecl:
                        v = r_data[extra_col]
                        if pd.notna(v): style_cell(ws1, row, 10, float(v), '0.000')
                    elif 'fdr' in ecl or 'q_value' in ecl or 'qvalue' in ecl:
                        v = r_data[extra_col]
                        if pd.notna(v): style_cell(ws1, row, 13, float(v), '0.000')
                    elif 'hub' in ecl:
                        v = r_data[extra_col]
                        if pd.notna(v): style_cell(ws1, row, 15, str(v))
                    elif 'severity' in ecl or 'sev_assoc' in ecl:
                        v = r_data[extra_col]
                        if pd.notna(v): style_cell(ws1, row, 16, str(v))
                    elif 'lr_pair' in ecl or 'top_lr' in ecl:
                        v = r_data[extra_col]
                        if pd.notna(v): style_cell(ws1, row, 17, str(v))
                    elif 'lr_contrib' in ecl or 'contribution' in ecl:
                        v = r_data[extra_col]
                        if pd.notna(v): style_cell(ws1, row, 18, float(v), '0.0')
                
                row += 1
            
            replaced['S1'] = row - 5
            print(f"  ✅ S1 populated: {row-5} pairs")
            return True
    
    # Strategy B: Separate CS + CSp matrices
    if cs_csv and csp_csv:
        cs_df = pd.read_csv(cs_csv, index_col=0)
        csp_df = pd.read_csv(csp_csv, index_col=0)
        print(f"\n  S1 from matrices: CS {cs_df.shape}, CSp {csp_df.shape}")
        
        # Clear old data
        for r in range(5, ws1.max_row + 1):
            for c in range(1, 19):
                ws1.cell(row=r, column=c).value = None
        
        row = 5
        pair_id = 1
        for fb in cs_df.index:
            for im in cs_df.columns:
                cs_val = cs_df.loc[fb, im]
                csp_val = csp_df.loc[fb, im] if fb in csp_df.index and im in csp_df.columns else None
                
                style_cell(ws1, row, 1, pair_id)
                style_cell(ws1, row, 2, str(fb), align='left')
                style_cell(ws1, row, 3, str(im))
                if pd.notna(cs_val):
                    style_cell(ws1, row, 4, float(cs_val), '0.000')
                if csp_val is not None and pd.notna(csp_val):
                    style_cell(ws1, row, 5, float(csp_val), '0.000')
                
                row += 1
                pair_id += 1
        
        replaced['S1'] = pair_id - 1
        print(f"  ✅ S1 populated from matrices: {pair_id-1} pairs")
        return True
    
    print("  ⚠️  S1: No source found — keeping template data")
    return False

populate_s1()

# ═══════════════════════════════════════════════════════════════
# TABLE S2: L-R CO-EXPRESSION
# ═══════════════════════════════════════════════════════════════
ws2 = wb["Table S2 - LR Co-expression"]

def populate_s2():
    if not lr_csv:
        print("  ⚠️  S2: No source found — keeping template data")
        return False
    
    df = pd.read_csv(lr_csv)
    print(f"\n  S2 source: {lr_csv}")
    print(f"  Columns: {list(df.columns)}")
    print(f"  Shape: {df.shape}")
    
    # Map columns
    col_map = {}
    for col in df.columns:
        cl = col.lower().strip()
        if cl in ['lr_pair', 'pair', 'l_r', 'ligand_receptor']:
            col_map['pair'] = col
        elif cl in ['ligand']:
            col_map['ligand'] = col
        elif cl in ['receptor']:
            col_map['receptor'] = col
        elif cl in ['condition', 'group', 'state']:
            col_map['condition'] = col
        elif cl in ['spatial_r', 'rho', 'r', 'spearman_r', 'correlation']:
            col_map['rho'] = col
        elif cl in ['spatial_p', 'p', 'pvalue', 'p_value']:
            col_map['p'] = col
        elif 'coupling' in cl or 'primary' in cl:
            col_map['coupling'] = col
    
    print(f"  Mapped: {col_map}")
    
    if 'pair' not in col_map or 'rho' not in col_map:
        print("  ⚠️  S2: Cannot map essential columns — keeping template")
        return False
    
    # Check if data is long-form (one row per pair × condition) or wide
    if 'condition' in col_map:
        # Long form → pivot to wide
        conditions = df[col_map['condition']].unique()
        print(f"  Conditions found: {conditions}")
        
        # Get unique pairs
        pairs = df[col_map['pair']].unique()
        
        # Clear old data
        for r in range(5, ws2.max_row + 1):
            for c in range(1, 13):
                ws2.cell(row=r, column=c).value = None
        
        row = 5
        rank = 1
        
        # Compute mean |rho| for ranking
        pair_means = df.groupby(col_map['pair'])[col_map['rho']].apply(lambda x: x.abs().mean())
        pair_means = pair_means.sort_values(ascending=False)
        
        for pair_name in pair_means.index:
            sub = df[df[col_map['pair']] == pair_name]
            
            ligand = sub[col_map.get('ligand', col_map['pair'])].iloc[0] if 'ligand' in col_map else pair_name.split('–')[0] if '–' in pair_name else pair_name.split('-')[0] if '-' in pair_name else ''
            receptor = sub[col_map.get('receptor', col_map['pair'])].iloc[0] if 'receptor' in col_map else pair_name.split('–')[-1] if '–' in pair_name else pair_name.split('-')[-1] if '-' in pair_name else ''
            coupling = sub[col_map['coupling']].iloc[0] if 'coupling' in col_map else ''
            
            style_cell(ws2, row, 1, rank)
            style_cell(ws2, row, 2, str(pair_name), align='left')
            style_cell(ws2, row, 3, str(ligand))
            style_cell(ws2, row, 4, str(receptor))
            style_cell(ws2, row, 5, str(coupling))
            
            # Fill condition-specific values
            for _, cond_row in sub.iterrows():
                cond = str(cond_row[col_map['condition']]).lower()
                rho_val = cond_row[col_map['rho']]
                p_val = cond_row.get(col_map.get('p', ''), None)
                
                if 'health' in cond or 'control' in cond:
                    if pd.notna(rho_val): style_cell(ws2, row, 6, float(rho_val), '0.000')
                    if p_val is not None and pd.notna(p_val): style_cell(ws2, row, 7, float(p_val), '0.00E+00')
                elif 'non' in cond:
                    if pd.notna(rho_val): style_cell(ws2, row, 8, float(rho_val), '0.000')
                    if p_val is not None and pd.notna(p_val): style_cell(ws2, row, 9, float(p_val), '0.00E+00')
                elif 'lesion' in cond:
                    if pd.notna(rho_val): style_cell(ws2, row, 10, float(rho_val), '0.000')
                    if p_val is not None and pd.notna(p_val): style_cell(ws2, row, 11, float(p_val), '0.00E+00')
            
            # Δ = lesional - healthy
            h_val = ws2.cell(row=row, column=6).value
            l_val = ws2.cell(row=row, column=10).value
            if h_val is not None and l_val is not None:
                style_cell(ws2, row, 12, round(float(l_val) - float(h_val), 3), '0.000')
            
            row += 1
            rank += 1
        
        replaced['S2'] = rank - 1
        print(f"  ✅ S2 populated: {rank-1} L-R pairs")
        return True
    
    else:
        # Wide form — try direct mapping
        print("  ⚠️  S2: Wide format detected, attempting direct copy...")
        # Clear and repopulate
        for r in range(5, ws2.max_row + 1):
            for c in range(1, 13):
                ws2.cell(row=r, column=c).value = None
        
        row = 5
        for idx, r_data in df.iterrows():
            style_cell(ws2, row, 1, idx + 1)
            style_cell(ws2, row, 2, str(r_data[col_map['pair']]), align='left')
            if 'ligand' in col_map: style_cell(ws2, row, 3, str(r_data[col_map['ligand']]))
            if 'receptor' in col_map: style_cell(ws2, row, 4, str(r_data[col_map['receptor']]))
            style_cell(ws2, row, 10, float(r_data[col_map['rho']]), '0.000')
            if 'p' in col_map and pd.notna(r_data[col_map['p']]):
                style_cell(ws2, row, 11, float(r_data[col_map['p']]), '0.00E+00')
            row += 1
        
        replaced['S2'] = row - 5
        print(f"  ✅ S2 populated (wide): {row-5} pairs")
        return True

populate_s2()

# ═══════════════════════════════════════════════════════════════
# TABLE S3: SAMPLE METRICS
# ═══════════════════════════════════════════════════════════════
ws3 = wb["Table S3 - Sample Metrics"]

def populate_s3():
    src = sample_csv or gsd_csv
    if not src:
        print("  ⚠️  S3: No source found — keeping template data")
        return False
    
    df = pd.read_csv(src)
    print(f"\n  S3 source: {src}")
    print(f"  Columns: {list(df.columns)}")
    
    # Also load GSD if separate
    gsd_df = None
    if gsd_csv and gsd_csv != src:
        gsd_df = pd.read_csv(gsd_csv)
        print(f"  GSD source: {gsd_csv}, cols: {list(gsd_df.columns)}")
    
    # Map columns
    col_map = {}
    for col in df.columns:
        cl = col.lower().strip()
        if cl in ['sample', 'sample_id', 'id', 'biopsy']:
            col_map['sample'] = col
        elif cl in ['condition', 'group', 'state', 'type']:
            col_map['condition'] = col
        elif cl in ['pasi', 'pasi_score']:
            col_map['pasi'] = col
        elif 'severity' in cl or 'class' in cl:
            col_map['severity'] = col
        elif 'mean_cs' in cl or 'cs_mean' in cl:
            col_map['mean_cs'] = col
        elif 'mast' in cl and 'cs' in cl:
            col_map['f5_mast'] = col
        elif 'trm' in cl and 'cs' in cl:
            col_map['f2_trm'] = col
        elif 'complex' in cl:
            col_map['complexity'] = col
        elif cl in ['pc1', 'pc1_score']:
            col_map['pc1'] = col
        elif cl in ['pc2', 'pc2_score']:
            col_map['pc2'] = col
        elif 'basin' in cl:
            col_map['basin'] = col
        elif 'depth' in cl:
            col_map['depth'] = col
        elif 'coher' in cl:
            col_map['coherence'] = col
        elif 'escape' in cl:
            col_map['escape'] = col
        elif 'trans' in cl and 'prob' in cl:
            col_map['transition'] = col
        elif 'dim' in cl and 'attractor' in cl:
            col_map['att_dim'] = col
    
    # Also scan columns with CS_ prefix for condition-specific
    cs_cols = {c: c for c in df.columns if c.startswith('CS_') or c.startswith('cs_')}
    
    print(f"  Mapped: {col_map}")
    
    if 'sample' not in col_map:
        print("  ⚠️  S3: No sample ID column — keeping template")
        return False
    
    # Clear old data
    for r in range(5, ws3.max_row + 1):
        for c in range(1, 17):
            ws3.cell(row=r, column=c).value = None
    
    row = 5
    s3_col_targets = {
        'sample': (1, None, 'left'),
        'condition': (2, None, 'center'),
        'pasi': (3, '0.0', 'center'),
        'severity': (4, None, 'center'),
        'mean_cs': (5, '0.000', 'center'),
        'f5_mast': (6, '0.000', 'center'),
        'f2_trm': (7, '0.000', 'center'),
        'complexity': (8, None, 'center'),
        'pc1': (9, '0.00', 'center'),
        'pc2': (10, '0.00', 'center'),
        'basin': (11, None, 'center'),
        'depth': (12, '0.000', 'center'),
        'coherence': (13, '0.000', 'center'),
        'escape': (14, '0.000', 'center'),
        'transition': (15, '0.000', 'center'),
        'att_dim': (16, None, 'center'),
    }
    
    for idx, r_data in df.iterrows():
        for key, (col_idx, fmt, align) in s3_col_targets.items():
            if key in col_map:
                val = r_data.get(col_map[key], None)
                if val is not None and pd.notna(val):
                    if fmt and isinstance(val, (int, float, np.floating, np.integer)):
                        style_cell(ws3, row, col_idx, float(val), fmt, align)
                    else:
                        style_cell(ws3, row, col_idx, str(val) if not isinstance(val, (int, float)) else val, fmt, align)
        row += 1
    
    # Merge GSD data if from separate file
    if gsd_df is not None:
        print(f"  Merging GSD data...")
        gsd_sample_col = None
        for c in gsd_df.columns:
            if c.lower() in ['sample', 'sample_id', 'id']:
                gsd_sample_col = c
                break
        
        if gsd_sample_col:
            for r in range(5, row):
                sid = ws3.cell(row=r, column=1).value
                match = gsd_df[gsd_df[gsd_sample_col] == sid]
                if len(match) > 0:
                    m = match.iloc[0]
                    for c in gsd_df.columns:
                        cl = c.lower()
                        if 'pc1' in cl and ws3.cell(row=r, column=9).value is None:
                            style_cell(ws3, row, 9, float(m[c]), '0.00')
                        elif 'pc2' in cl and ws3.cell(row=r, column=10).value is None:
                            style_cell(ws3, row, 10, float(m[c]), '0.00')
                        elif 'depth' in cl:
                            style_cell(ws3, row, 12, float(m[c]), '0.000')
                        elif 'coher' in cl:
                            style_cell(ws3, row, 13, float(m[c]), '0.000')
                        elif 'escape' in cl:
                            style_cell(ws3, row, 14, float(m[c]), '0.000')
    
    replaced['S3'] = row - 5
    print(f"  ✅ S3 populated: {row-5} samples")
    return True

populate_s3()

# ═══════════════════════════════════════════════════════════════
# TABLE S4: Keep as-is (curated)
# ═══════════════════════════════════════════════════════════════
print(f"\n  ℹ️  S4: Immune signatures — kept as curated template")
replaced['S4'] = 'template (curated)'

# ═══════════════════════════════════════════════════════════════
# SAVE & REPORT
# ═══════════════════════════════════════════════════════════════
os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
wb.save(OUTPUT)

print(f"\n{'='*60}")
print(f"FINAL REPORT")
print(f"{'='*60}")
print(f"Output: {OUTPUT}")
print(f"\nTable status:")
for table, count in replaced.items():
    status = f"✅ {count} rows replaced" if isinstance(count, int) else f"ℹ️  {count}"
    print(f"  {table}: {status}")

missing = []
if 'S1' not in replaced: missing.append('S1 (SICAI results)')
if 'S2' not in replaced: missing.append('S2 (L-R co-expression)')
if 'S3' not in replaced: missing.append('S3 (sample metrics)')

if missing:
    print(f"\n⚠️  Tables still using template data: {', '.join(missing)}")
    print(f"   These need actual CSVs from your pipeline.")
    print(f"   Expected CSV names and search directories:")
    print(f"   S1: sicai_results.csv or cs_matrix.csv + csp_matrix.csv")
    print(f"   S2: Table_S2_LR_spatial.csv")
    print(f"   S3: sample_metrics.csv or gsd_scores.csv")
    print(f"   Searched in: {RESULTS_DIRS}")
else:
    print(f"\n🎉 ALL TABLES POPULATED WITH REAL DATA!")

print(f"\nFile size: {os.path.getsize(OUTPUT)/1024:.0f} KB")
